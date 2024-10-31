from __future__ import print_function

import io
import os
import time

import dotenv
import openpyxl
from apiclient import discovery
from httplib2 import Http
from oauth2client import client
from oauth2client import file
from oauth2client import tools

dotenv.load_dotenv()

# Установка прав доступа
SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly']
all_data = []

name_spreadsheet = os.getenv('SPREAD_SHEET')


def get_credentials():
    """
    Выдача прав на доступ к google disk
    :return: права доступа
    """
    store = file.Storage('secret_data/storage.json')
    creds = store.get()
    # Если нет прав или они не валидны
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('secret_data/client_secret.json', SCOPES)
        creds = tools.run_flow(flow, store)
    return creds


def get_spreadsheet(name: str) -> None:
    """
    Загрузка документа в формате xlsx
    :param name: название документа
    :return: None
    """
    # Удаление прошлого файла
    if os.path.exists('file.xlsx'):
        os.remove('file.xlsx')
    # Получение прав
    credentials = get_credentials()
    service = discovery.build('drive', 'v3', http=credentials.authorize(Http()))

    # Сортировка по название таблицы/папки
    results = service.files().list(
        pageSize=10, q=f'name contains "{name}"', fields="files(id)").execute()
    # Получение id таблицы
    file_id = results['files'][0]['id']
    # Экспорт таблицы
    request = service.files().export(
        fileId=file_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet").execute()
    # Создание таблицы в папке проекта в формате xlsx
    with io.FileIO(os.path.join('.', 'file.xlsx'), 'wb') as file_write:
        file_write.write(request)


def get_data_from_spreadsheet(path: str) -> list[dict[str, str | None]]:
    """
    Получение данных из таблицы
    :param path: путь до таблицы
    :return: list[dict[str, str | None]]
    """
    # Загрузка таблицы
    workbook = openpyxl.load_workbook(path)
    # Активация таблицы
    sheet = workbook.active
    result = []
    # Прохождение по всем рядам и столбцам
    for i in range(5, 200):
        sheet_info = sheet.cell(row=i, column=2)
        # Если ячейка пуста
        if sheet_info.value is None:
            continue
        # Добавление в список: название документа, ссылку на него(если она есть), примечание и предложение(если они есть)
        result.append({
            'document': sheet_info.value,
            'link': sheet_info.hyperlink.target if sheet_info.hyperlink is not None else None,
            'note': sheet.cell(row=i, column=3).value if sheet.cell(row=i, column=3).hyperlink is None else sheet.cell(
                row=i, column=3).hyperlink.target,
            'offers': sheet.cell(row=i, column=4).value if sheet.cell(row=i,
                                                                      column=4).hyperlink is None else sheet.cell(row=i,
                                                                                                                  column=4).hyperlink.target,
        })
    return result


def main() -> None:
    """
    Загрузка документа и получение списка документов
    :return: None
    """
    global all_data
    # Скачивание таблицы
    get_spreadsheet(name_spreadsheet)
    # Загрузка из нее данных
    all_data = get_data_from_spreadsheet('file.xlsx')


# Запуск функции
main()
