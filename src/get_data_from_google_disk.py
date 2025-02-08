from __future__ import print_function

import asyncio
import io
import os
import time
from typing import Any

import dotenv
import openpyxl
from apiclient import discovery
from oauth2client import client
from oauth2client import file
from oauth2client import tools

dotenv.load_dotenv()

# Установка прав доступа
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

# Данные вперемешку
all_data = []
# Данные структурированные по папкам
all_data_with_folder = []
# Данные о сотрудниках
data_by_employees = []
# Месяцы по которым есть данные по сотрудникам
months = []

spreadsheet_KPI = os.getenv('SPREAD_SHEET_1')
spreadsheet_gmails = os.getenv('SPREAD_SHEET_2')
spreadsheet_objects = os.getenv('SPREAD_SHEET_3')

secret_path = 'C:/Users/new/PycharmProjects/telegram-bot/secret_data/client_secret.json'
storage_path = 'C:/Users/new/PycharmProjects/telegram-bot/secret_data/storage.json'
storage_path_amvera = '/app/secret_data/storage.json'
secret_path_amvera = '/app/secret_data/client_secret.json'

file_path_1 = '/data/file1.xlsx'
file_path_2 = '/data/file2.xlsx'
file_path_3 = '/data/file3.xlsx'


def get_credentials():
    """
    Выдача прав на доступ к google disk
    :return: права доступа
    """
    store = file.Storage(storage_path_amvera)
    creds = store.get()
    # Если нет прав или они не валидны
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(secret_path_amvera, SCOPES)
        creds = tools.run_flow(flow, store)
    return creds


# Получение прав
credentials = get_credentials()
service = discovery.build('drive', 'v3', credentials=credentials)
# Счетчик
counter = 0
# Ненужные папки
not_need_folders = []  # 'Фотофиксация выявленных дефектов.'
# Папки
folders = []
# Файлы
files = []


async def get_files_in_folder(service, folder_id: str) -> None:
    """
    Получение всех файлов в папке
    :param service:
    :param folder_id: str
    :return: None
    """
    global counter
    global files
    global folders
    response = (
        service.files()
        .list(
            fields="files(id, name, mimeType)",
            q=f"'{folder_id}' in parents",
              # f" and not name contains '.jpg' "
              # f"and not name contains '.jpeg'",
            pageSize=1000,
        )
        .execute()
    )
    # Обработка данных
    for file in response.get("files", []):
        # Если файл является папкой
        if file["mimeType"] == 'application/vnd.google-apps.folder':
            counter += 1
            if file["name"] in not_need_folders:
                continue
            all_data.append({'name': f'{file["name"]} (папка)',
                             'data': f'{file["id"]}',
                             'link': f'https://drive.google.com/drive/u/0/folders/{file["id"]}',
                             })
            folders.append({'name': f'{file["name"]} (папка)',
                            'data': f'{file["id"]}',
                            'link': f'https://drive.google.com/drive/u/0/folders/{file["id"]}',
                            'in_folder': f'{folder_id}',
                            'documents': []
                            })
            # Вход в рекурсию, для получения данных внутри этой папки
            await get_files_in_folder(service, file["id"])
        else:
            # Если файл является документом
            if file["mimeType"] in ['application/vnd.google-apps.document',
                                    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                    'application/msword']:
                link = f'https://docs.google.com/document/d/{file["id"]}'
            # Если файл является таблицей
            elif file["mimeType"] in ['application/vnd.google-apps.spreadsheet',
                                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                      'application/vnd.ms-excel']:
                link = f'https://docs.google.com/spreadsheets/d/{file["id"]}'
            # Если файл является презентацией
            elif file["mimeType"] in ['application/vnd.google-apps.presentation',
                                      'application/vnd.openxmlformats-officedocument.presentationml.presentation']:
                link = f'https://docs.google.com/presentation/d/{file["id"]}'
            # Если файл другого разрешения
            elif file["mimeType"] in ['image/tiff', 'application/pdf', 'image/png',
                                      'text/plain', 'text/html', 'application/zip',
                                      'application/vnd.ms-outlook', 'video/mp4', 'application/octet-stream',
                                      'image/jpeg']:
                link = f'https://docs.google.com/file/d/{file["id"]}'
            else:
                continue

            counter += 1
            if file["mimeType"] not in ['image/jpeg', 'video/mp4', 'image/png']:
                all_data.append({
                    'name': f'{file["name"]}',
                    'data': f'{file['id']}',
                    'link': link,
                })
            files.append(
                {'name': f'{file["name"]}',
                 'data': f'{file["id"]}',
                 'link': link,
                 'in_folder': f'{folder_id}',
                 'documents': []
                 })

    return None


async def get_files_and_folders(service, folders) -> list[dict[str, str | list]]:
    global files
    # Обработка файлов
    for folder2 in folders:
        for folder in folders[::-1]:
            for file in files:
                if folder['data'] == file['in_folder']:
                    folder['documents'].append(file)
                    files.remove(file)
    # Обработка папок
    for folder2 in folders[::-1]:
        if folder2['documents'] is None:
            await get_files_and_folders(service, folder2['documents'])
        for folder in folders[::-1]:
            if folder2['data'] == folder['in_folder']:
                folder2['documents'].append(folder)
                folders.remove(folder)
    return folders + files


async def get_spreadsheet(name: str, service, path: str) -> None:
    """
    Загрузка документа в формате xlsx
    :param path: путь, где сохраняется файл
    :param service:
    :param name: название документа
    :return: None
    """
    # Удаление прошлого файла
    if os.path.exists(path):
        os.remove(path)
    # Сортировка по название таблицы/папки
    results = service.files().list(
        pageSize=10,
        q=f'name contains "{name}"',
        fields="files(id, name, mimeType)"
    ).execute()
    # Получение id таблицы
    file_id = results['files'][0]['id']
    mimetype = results['files'][0]['mimeType']
    # Скачивание таблицы
    download_excel_file(file_id, mimetype, service, path)


def download_excel_file(file_id, mimetype, service, path):
    # Запрос на скачивание
    if mimetype == 'application/vnd.google-apps.spreadsheet':
        request = service.files().export(fileId=file_id,
                                         mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet').execute()

        with open(path, 'wb') as f:
            f.write(request)
    else:
        request = service.files().get_media(fileId=file_id).execute()

        with io.FileIO(os.path.join(path), 'wb') as file_write:
            file_write.write(request)


async def get_data_from_spreadsheet_KPI(month: int, service=service, path: str = file_path_1) -> list[dict[str, list[Any] | Any]]:
    """
    Получение данных из таблицы
    :param service:
    :param month: индекс страницы с месяцем за который идут данные
    :param path: путь до таблицы
    :return: list[dict[str, list[Any] | Any]]
    """
    await get_spreadsheet(spreadsheet_KPI, service, path)
    global months
    # Загрузка таблицы
    workbook = openpyxl.load_workbook(path)
    # Активация таблицы
    sheet = workbook.worksheets[month]
    months = [workbook.worksheets[i].title for i in range(len(workbook.worksheets))]
    employees = []
    # Прохождение по всем рядам и столбцам
    for col in range(10, 23):
        if employee_name := sheet.cell(row=8, column=col).value:
            employees.append({
                'name': employee_name,
                'info': [
                    sheet.cell(row=row, column=col).value
                    for row in range(10, 26)
                ]
            })

    return employees


async def get_data_from_spreadsheet_employees_gmail(service=service, path: str = file_path_2) -> list[dict[str, list[Any] | Any]]:
    """
    Получение данных из таблицы с gmail почтами сотрудников
    :param service:
    :param path: путь до таблицы
    :return: list[dict[str, list[Any] | Any]]
    """
    await get_spreadsheet(spreadsheet_gmails, service, path)
    # Загрузка таблицы
    workbook = openpyxl.load_workbook(path)
    # Активация таблицы
    sheet = workbook.active
    employees_gmail = []
    # Прохождение по всем рядам и столбцам
    for col in range(1, 2):
        for row in range(2, 20):
            if sheet.cell(row, 1).value and sheet.cell(row, 2).value:  # проверка на пустые ячейки
                employees_gmail.append({
                    'name': sheet.cell(row, 1).value,
                    'gmail': sheet.cell(row, 2).value,
                })
    return employees_gmail


async def get_data_from_spreadsheet_employees_objects(service=service, path: str = file_path_3) -> list[
    dict[str, list[Any | None] | Any]]:
    """
    Получение данных из таблицы с gmail почтами сотрудников
    :param service:
    :param path: путь до таблицы
    :return: list[dict[str, list[Any] | Any]]
    """
    await get_spreadsheet(spreadsheet_objects, service, path)
    # Загрузка таблицы
    workbook = openpyxl.load_workbook(path)
    # Активация таблицы
    sheet = workbook.active
    employees_objects = []
    # Прохождение по всем рядам и столбцам
    rows = []

    for row in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=3, values_only=True):
        rows.append(row)

    employees_objects_dict = {
        'name': '',
        'city': '',
        'objects': [],
    }
    for i in rows:
        if i[2]:
            employees_objects.append(employees_objects_dict.copy())
            employees_objects_dict['objects'] = []
            employees_objects_dict['name'] = i[2]
        if i[0]:
            employees_objects_dict['city'] = i[0]
        if i[1]:
            employees_objects_dict['objects'].append(i[1])

    return employees_objects[1:]


async def get_data_about_gmails_and_employees(month):
    global data_by_employees

    data_by_employees = await get_data_from_spreadsheet_KPI(month)
    gmails_by_employees = await get_data_from_spreadsheet_employees_gmail()
    objects_by_employees = await get_data_from_spreadsheet_employees_objects()

    for i in range(len(data_by_employees)):
        for j in range(len(gmails_by_employees)):
            if data_by_employees[i]['name'].split()[0] == gmails_by_employees[j]['name'].split()[0]:
                data_by_employees[i]['gmail'] = gmails_by_employees[j]['gmail']
        for j in range(len(objects_by_employees)):
            if data_by_employees[i]['name'].split()[0] == objects_by_employees[j]['name'].split()[0]:
                data_by_employees[i]['city'] = objects_by_employees[j]['city']
                data_by_employees[i]['objects'] = objects_by_employees[j]['objects']

    data_by_employees.sort(key=lambda x: x['name'])


async def main() -> None:
    """
    Получение данных
    :return: None
    """
    global all_data
    global all_data_with_folder
    global data_by_employees

    start = time.time()

    # Скачивание и получение данных из гугл таблиц
    await get_data_from_spreadsheet_employees_objects()
    await get_data_about_gmails_and_employees(len(months) - 1)

    # Получение списка документов из папки и ее подпапок
    await get_files_in_folder(service, '1NgZAEj6R507Qw8T1jS-2La5rrSNJqfXS')
    all_data_with_folder = await get_files_and_folders(service, folders)
    # Сортировка списков по названию документа и фамилиям
    all_data_with_folder.sort(key=lambda x: x['name'])
    all_data.sort(key=lambda x: x['name'])

    end = time.time()
    print(f'Время выполнения: {end - start} секунд')

asyncio.run(main())
