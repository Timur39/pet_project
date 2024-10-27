from __future__ import print_function

import io
import os

import openpyxl
from apiclient import discovery
from httplib2 import Http
from oauth2client import client
from oauth2client import file
from oauth2client import tools

SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly']


def get_credentials():
    store = file.Storage('../secret_data/storage.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('../secret_data/client_secret.json', SCOPES)
        creds = tools.run_flow(flow, store)
    return creds


def get_spreadsheet(name: str):
    credentials = get_credentials()
    service = discovery.build('drive', 'v3', http=credentials.authorize(Http()))
    results = service.files().list(
        pageSize=10, q=f'name contains "{name}"', fields="files(id)").execute()
    file_id = results['files'][0]['id']
    request = service.files().export(
        fileId=file_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet").execute()

    with io.FileIO(os.path.join('C:/Users/new/PycharmProjects/Pet_project', 'file.xlsx'), 'wb') as file_write:
        file_write.write(request)


def get_data_from_spreadsheet(path: str = 'C:/Users/new/PycharmProjects/Pet_project/file.xlsx'):
    # Определить переменную для загрузки книги
    workbook = openpyxl.load_workbook(path)
    # Определите переменную для чтения активного листа
    sheet = workbook.active
    result = []
    for i in range(5, 200):
        sheet_info = sheet.cell(row=i, column=2)
        if sheet_info.value is None:
            continue
        result.append({
            'document': sheet_info.value,
            'link': sheet_info.hyperlink.target if sheet_info.hyperlink is not None else None,
            'note': sheet.cell(row=i, column=3).value if sheet.cell(row=i, column=3).hyperlink is None else sheet.cell(row=i, column=3).hyperlink.target,
            'offers': sheet.cell(row=i, column=4).value if sheet.cell(row=i, column=4).hyperlink is None else sheet.cell(row=i, column=4).hyperlink.target,
        })
    return result


def main():
    get_spreadsheet('Сводная информация Отдел гарантийного сервиса')
    print(get_data_from_spreadsheet('C:/Users/new/PycharmProjects/Pet_project/file.xlsx'))


if __name__ == '__main__':
    main()
