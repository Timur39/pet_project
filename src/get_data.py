from __future__ import print_function

import io
import os

import dotenv
import openpyxl
from apiclient import discovery
from oauth2client import client
from oauth2client import file
from oauth2client import tools

dotenv.load_dotenv()

# Установка прав доступа
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

all_data = []
all_data_no_folders = []

name_spreadsheet = os.getenv('SPREAD_SHEET')


def get_credentials():
    """
    Выдача прав на доступ к google disk
    :return: права доступа
    """
    store = file.Storage('/app/secret_data/storage.json')
    creds = store.get()
    # Если нет прав или они не валидны
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(
            '/app/secret_data/client_secret.json', SCOPES)
        creds = tools.run_flow(flow, store)
    return creds


def get_folders_and_files():
    # Получение прав
    credentials = get_credentials()
    service = discovery.build('drive', 'v3', credentials=credentials)
    # Сортировка по название папки/файла
    response_for_folder = response = (
        service.files()
        .list(
            fields="files(id)",
            q="name = 'Гарантийные обязательства'",
            pageSize=1,
        )
        .execute()
    )
    response = (
        service.files()
        .list(
            fields="nextPageToken, files(id, name, mimeType)",
            q=f"mimeType = 'application/vnd.google-apps.folder' and '{response_for_folder['files'][0]['id']}' in parents",
            pageSize=1000,
        )
        .execute()
    )
    all_data_no_folders_link = [data['link'] for data in all_data_no_folders]
    result = []
    state = True
    for file in response.get("files", []):
        for doc_link in all_data_no_folders_link:
            if doc_link is None:
                continue
            elif f'{file.get("id")}' in doc_link:
                state = False
        if state:
            result.append({'document': f'{file.get("name")} (папка)',
                           'link': f'https://drive.google.com/drive/u/0/folders/{file.get("id")}',
                           'note': None,
                           'offers': None
                           })
    return result


def get_spreadsheet(name: str) -> None:
    """
    Загрузка документа в формате xlsx
    :param name: название документа
    :return: None
    """
    # Удаление прошлого файла
    if os.path.exists('/app/file.xlsx'):
        os.remove('/app/file.xlsx')
    # Получение прав
    credentials = get_credentials()
    service = discovery.build('drive', 'v3', credentials=credentials)
    # Сортировка по название таблицы/папки
    results = service.files().list(
        pageSize=10, q=f'name contains "{name}"', fields="files(id)").execute()
    # Получение id таблицы
    file_id = results['files'][0]['id']
    # Экспорт таблицы
    request = service.files().export(
        fileId=file_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    request = request.execute()
    # Создание таблицы в папке проекта в формате xlsx
    with io.FileIO(os.path.join('/app', 'file.xlsx'), 'wb') as file_write:
        file_write.write(request)


def get_data_from_spreadsheet(path: str) -> list[dict[str, str | None]]:
    """
    Получение данных из таблицы
    :param path: путь до таблицы
    :return: list[dict[str, str | None]]
    """
    # Загрузка таблицы
    workbook = openpyxl.load_workbook(path)
    # Активация таблицы
    sheet = workbook.active
    result = []
    # Прохождение по всем рядам и столбцам
    for i in range(5, 200):
        sheet_info = sheet.cell(row=i, column=2)
        # Если ячейка пуста
        if sheet_info.value is None:
            continue
        # Добавление в список: название документа, ссылку на него(если она есть), примечание и предложение(если они есть)
        string = ''
        if 'folder' in sheet_info.hyperlink.target if sheet_info.hyperlink else '':
            string = '(папка)'
        elif 'folder' in sheet.cell(row=i, column=3).hyperlink.target if sheet.cell(row=i, column=3).hyperlink else '':
            string = '(папка)'
        elif 'folder' in sheet.cell(row=i, column=4).hyperlink.target if sheet.cell(row=i, column=4).hyperlink else '':
            string = '(папка)'

        result.append({
            'document': f'{sheet_info.value} {string}',
            'link': sheet_info.hyperlink.target if sheet_info.hyperlink else None,
            'note': sheet.cell(row=i, column=3).value if sheet.cell(row=i, column=3).hyperlink is None else sheet.cell(
                row=i, column=3).hyperlink.target,
            'offers': sheet.cell(row=i, column=4).value if sheet.cell(row=i,
                                                                      column=4).hyperlink is None else sheet.cell(row=i,
                                                                                                                  column=4).hyperlink.target,
        })
    return result


def main() -> None:
    """
    Загрузка документа и получение списка документов
    :return: None
    """
    global all_data
    global all_data_no_folders
    # Скачивание таблицы
    get_spreadsheet(name_spreadsheet)
    # Загрузка из нее данных
    all_data_no_folders = get_data_from_spreadsheet('/app/file.xlsx')
    all_data = get_data_from_spreadsheet('/app/file.xlsx') + get_folders_and_files()
    # Сортировка списка по названию документа
    all_data.sort(key=lambda x: x['document'])


main()
